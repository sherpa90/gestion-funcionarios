from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from permisos.models import SolicitudPermiso
from licencias.models import LicenciaMedica
from users.models import CustomUser
from core.services import BusinessDayCalculator
import openpyxl
from datetime import datetime

class ReportesView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista unificada y minimalista de reportes"""
    template_name = 'reportes/reportes.html'

    def test_func(self):
        # Acceso para Director, Secretaria, Admin y Directivos
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener parámetros de filtro
        search = self.request.GET.get('search', '')
        year = self.request.GET.get('year', '')
        fecha_inicio = self.request.GET.get('fecha_inicio', '')
        fecha_fin = self.request.GET.get('fecha_fin', '')
        sort_by = self.request.GET.get('sort', 'name')
        
        # Base queryset: incluir todos los funcionarios del sistema
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'])
        
        # Filtro de búsqueda por nombre o RUN
        if search:
            funcionarios = funcionarios.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(run__icontains=search)
            )
        
        # Preparar datos de cada funcionario
        empleados_data = []
        for funcionario in funcionarios:
            # Obtener permisos aprobados
            permisos = SolicitudPermiso.objects.filter(
                usuario=funcionario,
                estado='APROBADO'
            )
            
            # Aplicar filtros de fecha/año
            if year:
                permisos = permisos.filter(fecha_inicio__year=year)
            if fecha_inicio:
                permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
            
            dias_usados = permisos.aggregate(Sum('dias_solicitados'))['dias_solicitados__sum'] or 0
            
            # Obtener licencias médicas
            licencias = LicenciaMedica.objects.filter(usuario=funcionario)
            
            if year:
                licencias = licencias.filter(fecha_inicio__year=year)
            if fecha_inicio:
                licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                licencias = licencias.filter(fecha_inicio__lte=fecha_fin)
            
            total_licencias = licencias.count()
            dias_licencias = licencias.aggregate(Sum('dias'))['dias__sum'] or 0
            
            empleados_data.append({
                'funcionario': funcionario,
                'dias_disponibles': funcionario.dias_disponibles,
                'dias_usados': dias_usados,
                'total_licencias': total_licencias,
                'dias_licencias': dias_licencias,
            })
        
        # Aplicar ordenamiento
        if sort_by == 'name':
            empleados_data.sort(key=lambda x: (x['funcionario'].last_name, x['funcionario'].first_name))
        elif sort_by == 'name_desc':
            empleados_data.sort(key=lambda x: (x['funcionario'].last_name, x['funcionario'].first_name), reverse=True)
        elif sort_by == 'dias':
            empleados_data.sort(key=lambda x: x['dias_disponibles'], reverse=True)
        elif sort_by == 'dias_asc':
            empleados_data.sort(key=lambda x: x['dias_disponibles'])
        elif sort_by == 'dias_usados':
            empleados_data.sort(key=lambda x: x['dias_usados'], reverse=True)
        elif sort_by == 'dias_usados_asc':
            empleados_data.sort(key=lambda x: x['dias_usados'])
        elif sort_by == 'licencias':
            empleados_data.sort(key=lambda x: x['total_licencias'], reverse=True)
        elif sort_by == 'licencias_asc':
            empleados_data.sort(key=lambda x: x['total_licencias'])
        elif sort_by == 'dias_licencias':
            empleados_data.sort(key=lambda x: x['dias_licencias'], reverse=True)
        elif sort_by == 'dias_licencias_asc':
            empleados_data.sort(key=lambda x: x['dias_licencias'])
        
        context['empleados_data'] = empleados_data # Kept original key as per instruction, but note the new sorting keys might expect 'nombre' and 'dias_licencia'
        context['filtros'] = {
            'search': search,
            'year': year,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
        
        # Años disponibles para filtro
        from datetime import datetime
        permisos_years = set(SolicitudPermiso.objects.dates('fecha_inicio', 'year').values_list('fecha_inicio', flat=True))
        licencias_years = set(LicenciaMedica.objects.dates('fecha_inicio', 'year').values_list('fecha_inicio', flat=True))
        all_years = sorted(set([d.year for d in permisos_years] + [d.year for d in licencias_years]), reverse=True)
        context['years'] = all_years if all_years else [datetime.now().year]
        context['current_sort'] = sort_by
        
        return context


class PDFIndividualView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generar PDF de un solo empleado"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request, usuario_id):
        try:
            funcionario = CustomUser.objects.get(id=usuario_id)
        except CustomUser.DoesNotExist:
            return HttpResponse('Funcionario no encontrado', status=404)
        
        # Obtener parámetros de filtro
        year = request.GET.get('year', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        
        # Permisos
        permisos = SolicitudPermiso.objects.filter(usuario=funcionario, estado='APROBADO')
        if year:
            permisos = permisos.filter(fecha_inicio__year=year)
        if fecha_inicio:
            permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
        if fecha_fin:
            permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
        permisos = permisos.order_by('-fecha_inicio')
        
        # Licencias
        licencias = LicenciaMedica.objects.filter(usuario=funcionario)
        if year:
            licencias = licencias.filter(fecha_inicio__year=year)
        if fecha_inicio:
            licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
        if fecha_fin:
            licencias = licencias.filter(fecha_inicio__lte=fecha_fin)
        licencias = licencias.order_by('-fecha_inicio')
        
        # Prepare data for template
        from datetime import timedelta
        
        permisos_data = []
        for p in permisos:
            permisos_data.append({
                'fecha_inicio': p.fecha_inicio,
                'fecha_fin': p.fecha_termino,
                'dias_solicitados': p.dias_solicitados,
                'tipo': 'Permiso Administrativo',
                'motivo': p.observacion or "Administrativo"
            })
            
        licencias_data = []
        for l in licencias:
            fecha_fin_est = l.fecha_inicio + timedelta(days=l.dias - 1)
            licencias_data.append({
                'fecha_inicio': l.fecha_inicio,
                'fecha_fin': fecha_fin_est,
                'dias': l.dias,
                'tipo': 'Licencia Médica'
            })
        
        html_string = render_to_string('reportes/pdf_individual.html', {
            'funcionario': funcionario,
            'permisos': permisos_data,
            'licencias': licencias_data,
            'total_dias_usados': permisos.aggregate(Sum('dias_solicitados'))['dias_solicitados__sum'] or 0,
            'total_dias_licencias': licencias.aggregate(Sum('dias'))['dias__sum'] or 0,
            'year': year,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=reporte_{funcionario.run}.pdf'
        response.write(result)
        return response


class PDFColectivoView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generar PDF de todos los empleados filtrados"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        # Obtener parámetros de filtro
        search = request.GET.get('search', '')
        year = request.GET.get('year', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        
        # Filtrar funcionarios
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'])
        if search:
            funcionarios = funcionarios.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(run__icontains=search)
            )
        
        # Preparar datos
        empleados_data = []
        for funcionario in funcionarios.order_by('last_name', 'first_name'):
            permisos = SolicitudPermiso.objects.filter(usuario=funcionario, estado='APROBADO')
            if year:
                permisos = permisos.filter(fecha_inicio__year=year)
            if fecha_inicio:
                permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
            
            licencias = LicenciaMedica.objects.filter(usuario=funcionario)
            if year:
                licencias = licencias.filter(fecha_inicio__year=year)
            if fecha_inicio:
                licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                licencias = licencias.filter(fecha_inicio__lte=fecha_fin)
            
            empleados_data.append({
                'funcionario': funcionario,
                'dias_disponibles': funcionario.dias_disponibles,
                'dias_usados': permisos.aggregate(Sum('dias_solicitados'))['dias_solicitados__sum'] or 0,
                'total_licencias': licencias.count(),
                'dias_licencias': licencias.aggregate(Sum('dias'))['dias__sum'] or 0,
            })
        
        html_string = render_to_string('reportes/pdf_colectivo.html', {
            'empleados_data': empleados_data,
            'year': year,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'total_funcionarios': len(empleados_data),
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename=reporte_colectivo.pdf'
        response.write(result)
        return response


class ExportarExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Exportar reporte detallado a Excel"""
    
    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN', 'DIRECTIVO']

    def get(self, request):
        # Obtener parámetros de filtro
        search = request.GET.get('search', '')
        year = request.GET.get('year', '')
        fecha_inicio = request.GET.get('fecha_inicio', '')
        fecha_fin = request.GET.get('fecha_fin', '')
        
        # Crear libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Detallado"
        
        # Estilos
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        center_alignment = openpyxl.styles.Alignment(horizontal="center")
        
        # Encabezados
        headers = ['Funcionario', 'RUN', 'Tipo', 'Motivo', 'Fecha Inicio', 'Fecha Fin', 'Días', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            
        # Filtrar funcionarios
        funcionarios = CustomUser.objects.filter(role__in=['FUNCIONARIO', 'DIRECTOR', 'DIRECTIVO', 'SECRETARIA', 'ADMIN'])
        if search:
            funcionarios = funcionarios.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(run__icontains=search)
            )
            
        row_num = 2
        for funcionario in funcionarios.order_by('last_name', 'first_name'):
            # Permisos
            permisos = SolicitudPermiso.objects.filter(usuario=funcionario, estado='APROBADO')
            if year:
                permisos = permisos.filter(fecha_inicio__year=year)
            if fecha_inicio:
                permisos = permisos.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                permisos = permisos.filter(fecha_inicio__lte=fecha_fin)
                
            for permiso in permisos:
                ws.cell(row=row_num, column=1, value=funcionario.get_full_name())
                ws.cell(row=row_num, column=2, value=funcionario.run)
                ws.cell(row=row_num, column=3, value="Permiso Administrativo")
                ws.cell(row=row_num, column=4, value=permiso.observacion or "-")
                ws.cell(row=row_num, column=5, value=permiso.fecha_inicio.strftime('%d/%m/%Y'))
                ws.cell(row=row_num, column=6, value=permiso.fecha_termino.strftime('%d/%m/%Y') if permiso.fecha_termino else "-")
                ws.cell(row=row_num, column=7, value=permiso.dias_solicitados)
                ws.cell(row=row_num, column=8, value=permiso.get_estado_display())
                row_num += 1
                
            # Licencias
            licencias = LicenciaMedica.objects.filter(usuario=funcionario)
            if year:
                licencias = licencias.filter(fecha_inicio__year=year)
            if fecha_inicio:
                licencias = licencias.filter(fecha_inicio__gte=fecha_inicio)
            if fecha_fin:
                licencias = licencias.filter(fecha_inicio__lte=fecha_fin)
                
            for licencia in licencias:
                from datetime import timedelta
                # LicenciaMedica does not have fecha_fin, so we always calculate it
                fecha_fin_display = (licencia.fecha_inicio + timedelta(days=licencia.dias - 1)).strftime('%d/%m/%Y')
                
                ws.cell(row=row_num, column=1, value=funcionario.get_full_name())
                ws.cell(row=row_num, column=2, value=funcionario.run)
                ws.cell(row=row_num, column=3, value="Licencia Médica")
                ws.cell(row=row_num, column=4, value="-")
                ws.cell(row=row_num, column=5, value=licencia.fecha_inicio.strftime('%d/%m/%Y'))
                ws.cell(row=row_num, column=6, value=fecha_fin_display)
                ws.cell(row=row_num, column=7, value=licencia.dias)
                ws.cell(row=row_num, column=8, value="Presentada")
                row_num += 1
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        # Preparar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte_detallado.xlsx'
        wb.save(response)
        return response


class ReporteMensualDiasAdministrativosView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Generar PDF mensual de resumen de días administrativos"""

    def test_func(self):
        return self.request.user.role in ['DIRECTOR', 'SECRETARIA', 'ADMIN']

    def get(self, request):
        # Obtener parámetros del mes y año
        mes = request.GET.get('mes')
        anio = request.GET.get('anio')

        if not mes or not anio:
            # Usar mes y año actuales por defecto
            now = datetime.now()
            mes = str(now.month)
            anio = str(now.year)

        mes = int(mes)
        anio = int(anio)

        # Obtener todas las solicitudes aprobadas del mes
        solicitudes_mes = SolicitudPermiso.objects.filter(
            estado='APROBADO',
            fecha_inicio__year=anio,
            fecha_inicio__month=mes
        ).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')

        # Preparar datos para el reporte
        empleados_data = []
        for solicitud in solicitudes_mes:
            # Calcular fecha "desde" (fecha de inicio)
            fecha_desde = solicitud.fecha_inicio

            # Calcular fecha "hasta" usando BusinessDayCalculator
            fecha_hasta = BusinessDayCalculator.calculate_end_date(fecha_desde, solicitud.dias_solicitados)

            empleados_data.append({
                'run': solicitud.usuario.run,
                'nombre_completo': solicitud.usuario.get_full_name(),
                'establecimiento': 'Colegio Los Alerces',
                'dias_solicitados': solicitud.dias_solicitados,
                'dias_disponibles': solicitud.usuario.dias_disponibles,
                'fecha_desde': fecha_desde,
                'fecha_hasta': fecha_hasta,
            })

        # Nombres de meses en español
        meses = [
            '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
        ]

        # Obtener el director del sistema
        director = CustomUser.objects.filter(role='DIRECTOR').first()

        # Generar PDF
        html_string = render_to_string('reportes/reporte_mensual_dias_administrativos.html', {
            'empleados_data': empleados_data,
            'mes': meses[mes],
            'anio': anio,
            'mes_numero': mes,
            'establecimiento': 'Colegio Los Alerces de Puerto Montt',
            'director_nombre': director.get_full_name() if director else 'Director',
            'fecha_generacion': datetime.now(),
        })

        html = HTML(string=html_string)
        result = html.write_pdf()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=cuadro_resumen_dias_administrativos_{mes:02d}_{anio}.pdf'
        response.write(result)
        return response
