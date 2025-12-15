#!/usr/bin/env python
"""
Script para probar el formato del Excel de asistencia
"""
import os
import sys
import django

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
import re
from datetime import datetime, time

def test_excel_format():
    print("🧪 Probando formato del Excel de asistencia...")
    print("=" * 60)

    excel_file = 'templates/Asistentes_Nov.xlsx'

    if not os.path.exists(excel_file):
        print(f"❌ Archivo no encontrado: {excel_file}")
        return

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        print(f"📄 Archivo cargado. Filas totales: {ws.max_row}")

        # Mostrar primeras 10 filas
        print("\n📋 Primeras 10 filas:")
        print("Fila | RUT | Nombre | Horario")
        print("-" * 50)

        for row_num in range(1, min(11, ws.max_row + 1)):
            row = []
            for col in range(1, 4):  # Columnas A, B, C
                cell_value = ws.cell(row=row_num, column=col).value
                row.append(str(cell_value) if cell_value is not None else "")

            if row_num == 1:
                print("1 (Header) | {} | {} | {}".format(row[0], row[1], row[2]))
            else:
                print("{:2d} | {} | {} | {}".format(row_num, row[0], row[1], row[2][:20] if len(row[2]) > 20 else row[2]))

        # Probar parsing de algunas filas de datos
        print("\n🔍 Probando parsing de filas de datos:")

        for row_num in range(2, min(6, ws.max_row + 1)):
            rut_raw = ws.cell(row=row_num, column=1).value
            nombre = ws.cell(row=row_num, column=2).value
            horario_raw = ws.cell(row=row_num, column=3).value

            print(f"\n📝 Fila {row_num}:")
            print(f"   RUT: {rut_raw}")
            print(f"   Nombre: {nombre}")
            print(f"   Horario: {horario_raw}")

            # Probar el regex del código
            if horario_raw:
                horario_str = str(horario_raw).strip()
                fecha_hora_match = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})\s+(\d{1,2}):(\d{2})$', horario_str)

                if fecha_hora_match:
                    dia, mes, anio = fecha_hora_match.groups()[:3]
                    hora, minuto = fecha_hora_match.groups()[3:]

                    print("   ✅ Regex match:")
                    print(f"      Día: {dia}, Mes: {mes}, Año: {anio}")
                    print(f"      Hora: {hora}, Minuto: {minuto}")

                    # Intentar crear fecha
                    try:
                        fecha = datetime(int(anio), int(mes), int(dia)).date()
                        hora_entrada = time(int(hora), int(minuto))
                        print(f"   ✅ Fecha creada: {fecha}")
                        print(f"   ✅ Hora creada: {hora_entrada}")
                    except Exception as e:
                        print(f"   ❌ Error creando fecha/hora: {e}")
                else:
                    print("   ❌ Regex no coincide")
                    print(f"      Patrón esperado: DD-MM-YYYY HH:MM")
                    print(f"      Recibido: '{horario_str}'")

    except Exception as e:
        print(f"❌ Error procesando Excel: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    test_excel_format()