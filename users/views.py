from django.views.generic import ListView, CreateView, UpdateView, DeleteView, FormView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import CustomUser
from .forms import UserCreateForm, UserEditForm, BulkUserImportForm
import openpyxl
from openpyxl import Workbook
import random
import string
from io import BytesIO

from django.db.models import Q

class UserListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = CustomUser
    template_name = 'users/user_list.html'
    context_object_name = 'users'
    
    def test_func(self):
        return self.request.user.role in ['SECRETARIA', 'ADMIN']
    
    def get_queryset(self):
        # Orden por defecto: Nombre (A-Z) usando first_name primero
        queryset = CustomUser.objects.all().order_by('first_name', 'last_name')
        
        # Búsqueda
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) | 
                Q(last_name__icontains=search_query) | 
                Q(run__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        
        # Aplicar ordenamiento si se solicita
        sort_by = self.request.GET.get('sort', 'name')
        
        if sort_by == 'name':
            queryset = queryset.order_by('first_name', 'last_name')
        elif sort_by == 'name_desc':
            queryset = queryset.order_by('-first_name', '-last_name')
        elif sort_by == 'role':
            queryset = queryset.order_by('role', 'first_name')
        elif sort_by == 'role_desc':
            queryset = queryset.order_by('-role', 'first_name')
        elif sort_by == 'tipo':
            queryset = queryset.order_by('tipo_funcionario', 'first_name')
        elif sort_by == 'tipo_desc':
            queryset = queryset.order_by('-tipo_funcionario', 'first_name')
        elif sort_by == 'dias':
            queryset = queryset.order_by('-dias_disponibles')
        elif sort_by == 'dias_asc':
            queryset = queryset.order_by('dias_disponibles')
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'name')
        
        # Recuperar contraseñas de importación masiva y limpiar sesión
        if 'bulk_passwords' in self.request.session:
            context['bulk_passwords'] = self.request.session.pop('bulk_passwords')
            
        return context

class UserCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = CustomUser
    form_class = UserCreateForm
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return self.request.user.role in ['SECRETARIA', 'ADMIN']
    
    def form_valid(self, form):
        response = super().form_valid(form)
        password = self.object.generated_password
        messages.success(
            self.request, 
            f'Usuario creado exitosamente. Contraseña temporal: {password}'
        )
        return response

class UserUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    form_class = UserEditForm
    template_name = 'users/user_form.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return self.request.user.role in ['SECRETARIA', 'ADMIN']
    
    def form_valid(self, form):
        messages.success(self.request, 'Usuario actualizado exitosamente')
        return super().form_valid(form)

class UserDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = CustomUser
    template_name = 'users/user_confirm_delete.html'
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        return self.request.user.role in ['SECRETARIA', 'ADMIN']
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Usuario eliminado exitosamente')
        return super().delete(request, *args, **kwargs)


class BulkUserImportView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    """Vista para importación masiva de usuarios desde Excel"""
    template_name = 'users/bulk_import.html'
    form_class = BulkUserImportForm
    success_url = reverse_lazy('user_list')
    
    def test_func(self):
        # Solo admins y secretarias
        return self.request.user.role in ['ADMIN', 'SECRETARIA']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['download_template'] = True
        return context
    
    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']
        
        try:
            # Leer archivo Excel
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            created_users = []
            errors = []
            passwords = {}  # Almacenar contraseñas generadas
            
            # Procesar cada fila (empezar desde fila 2, asumiendo headers en fila 1)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue
                
                try:
                    run, first_name, last_name, email, role, tipo_funcionario, dias = row[:7]
                    
                    # Validaciones básicas
                    if not run or not first_name or not last_name:
                        errors.append(f"Fila {row_num}: Faltan datos obligatorios (RUN, nombre o apellido)")
                        continue
                    
                    # Validar email
                    if email and not email.endswith('@losalercespuertomontt.cl'):
                        errors.append(f"Fila {row_num}: Email debe ser @losalercespuertomontt.cl")
                        continue
                    
                    # Crear email si no existe
                    if not email:
                        email = f"{run}@losalercespuertomontt.cl"
                    
                    # Verificar si ya existe
                    if CustomUser.objects.filter(run=run).exists():
                        errors.append(f"Fila {row_num}: Usuario con RUN {run} ya existe")
                        continue
                    
                    # Generar contraseña
                    password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                    
                    # Crear usuario
                    user = CustomUser.objects.create_user(
                        run=run,
                        username=run,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        password=password,
                        role=role or 'FUNCIONARIO',
                        tipo_funcionario=tipo_funcionario or 'PLANTA',
                        dias_disponibles=dias or 6.0
                    )
                    
                    created_users.append(user)
                    passwords[run] = password
                    
                except Exception as e:
                    errors.append(f"Fila {row_num}: Error - {str(e)}")
            
            # Resultado
            if created_users:
                messages.success(
                    self.request,
                    f"Se crearon {len(created_users)} usuarios exitosamente"
                )
                
                # Guardar contraseñas para mostrar
                self.request.session['bulk_passwords'] = passwords
                
            if errors:
                for error in errors:
                    messages.warning(self.request, error)
            
            return redirect(self.success_url)
            
        except Exception as e:
            messages.error(self.request, f"Error al procesar archivo: {str(e)}")
            return self.form_invalid(form)


def download_template(request):
    """Generar y descargar plantilla Excel para importación"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Headers
    headers = ['RUN', 'Nombres', 'Apellidos', 'Email', 'Rol', 'Tipo Funcionario', 'Días Disponibles']
    ws.append(headers)
    
    # Ejemplo
    ws.append([
        '12345678-9',
        'Juan',
        'Pérez',
        'juan.perez@losalercespuertomontt.cl',
        'FUNCIONARIO',
        'PLANTA',
        6.0
    ])
    
    # Ajustar anchos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_usuarios.xlsx'
    
    return response
